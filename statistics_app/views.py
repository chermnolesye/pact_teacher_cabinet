from django.shortcuts import render
from core_app.models import (
    Text,
    Token,
    Error,
    Error,
	ErrorLevel,
    ErrorToken,
    ErrorTag,
    Group,
	Emotion,
    AcademicYear,
    TextType,
    Student,
    Sentence,
    User,
)

from collections import defaultdict
from statistics_app import dashboards
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Q
import json
import openpyxl
from openpyxl.styles import Font
import numpy as np
import scipy

def export_group_error_stats(request):
    group_id = request.GET.get('group')
    if not group_id:
        return HttpResponse("Группа не выбрана", status=400)

    try:
        group = Group.objects.get(idgroup=group_id)
    except Group.DoesNotExist:
        return HttpResponse("Группа не найдена", status=404)

    students = Student.objects.filter(idgroup=group_id).select_related("iduser").order_by("iduser__lastname")
    tags = {t.iderrortag: t.tagtext for t in ErrorTag.objects.all()}

    wb = openpyxl.Workbook()
    default_sheet = wb.active 

    added_sheets = False  

    for student in students:
        text_ids = Text.objects.filter(idstudent=student.idstudent).values_list('idtext', flat=True)

        error_counts_raw = (
            Error.objects
            .filter(errortoken__idtoken__idsentence__idtext__in=text_ids)
            .values("iderrortag")
            .annotate(count=Count("iderror"))
        )

        if not error_counts_raw:
            continue

        error_counts = []
        for error in error_counts_raw:
            tag_id = error["iderrortag"]
            tag_name = tags.get(tag_id, "Неизвестно")
            count = error["count"]
            error_counts.append((tag_name, count))
        error_counts.sort(key=lambda x: x[0])

        sheet_name = f"{student.iduser.lastname} {student.iduser.firstname} {student.iduser.middlename}"
        sheet_name = sheet_name[:31]
        ws = wb.create_sheet(title=sheet_name)

        ws.append(["Тэг", "Частота"])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for tag_name, count in error_counts:
            ws.append([tag_name, count])

        added_sheets = True

    if added_sheets:
        wb.remove(default_sheet)
    else:
        default_sheet.title = "Нет данных"
        default_sheet.append(["Нет ошибок у студентов в этой группе."])

    filename = f"{group.groupname} ({group.idayear.title}).xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def statistics_view(request):
    groups = (
        Group.objects.select_related("idayear")
        .all()
        .values("idgroup", "groupname", "idayear__title")
        .distinct()
    )
    group_data = [
        {
            "id": group["idgroup"],
            "name": group["groupname"],
            "year": group["idayear__title"],
        }
        for group in groups
    ]

    group_id = ""

    context = {"groups": group_data, "selected_group": group_id}

    return render(request, "statistics.html", context)


def error_stats(request):
    tags = ErrorTag.objects.all().values(
        "iderrortag",
        "tagtext",
        "tagtextrussian",
        "idtagparent",
    )

    # Считаем ошибки по тегу и уровню сложности
    error_counts = Error.objects.values("iderrortag", "iderrorlevel").annotate(
        count=Count("iderror")
    )

    # Карта: id тэга -> { уровень -> количество ошибок }
    error_count_map = defaultdict(lambda: defaultdict(int))
    for item in error_counts:
        tag_id = item["iderrortag"]
        level_id = item["iderrorlevel"]
        count = item["count"]
        error_count_map[tag_id][level_id] = count

    grouped_tags = defaultdict(list)

    for tag in tags:
        tag_id = tag["iderrortag"]
        tag_name = tag["tagtext"]
        parent_id = tag["idtagparent"]

        levels = error_count_map.get(tag_id, {})

        # Считаем активные уровни (где количество > 0)
        active_levels = [lvl for lvl, cnt in levels.items() if cnt > 0]
        num_active_levels = len(active_levels)

        # Определяем цвет
        if num_active_levels == 0 or num_active_levels == 1:
            color = "#ffffff"
        elif num_active_levels == 2:
            color = "#e8f0fe"
        else:
            color = "#cfe2ff"

        tag_info = {
            "id": tag_id,
            "nametag": tag_name,
            "color": color,
            "level1": levels.get(1, 0),
            "level2": levels.get(2, 0),
            "level3": levels.get(3, 0),
            "parent_id": parent_id,  # Добавляем информацию о родителе
        }

        if parent_id:
            parent_name = next(
                (t["tagtext"] for t in tags if t["iderrortag"] == parent_id), None
            )
            if parent_name:
                grouped_tags[parent_name].append(tag_info)
        else:
            grouped_tags[tag_name].append(tag_info)  # Родительские теги

    context = {"tags_error": dict(grouped_tags)}
    return render(request, "error_stats.html", context)


def chart_types_errors2(request):
    if request.method != "POST":
        levels = dashboards.get_levels()

        groups = list(
            Group.objects.values("groupname", "idayear")
            .distinct()
            .order_by("groupname")
        )

        courses = list(
            Group.objects.values("studycourse", "idayear")
            .filter(studycourse__gt=0)
            .distinct()
            .order_by("studycourse")
        )

        texts = list(
            Text.objects.values("header")
            .filter(errorcheckflag=True)
            .distinct()
            .order_by("header")
        )

        text_types = list(
            TextType.objects.filter(text__errorcheckflag=True)
            .distinct()
            .order_by("idtexttype")
            .values()
        )

        # Подсчёт ошибок по тегам
        data_count_errors = list(
            ErrorToken.objects.values(
                "iderror__iderrortag__iderrortag",
                "iderror__iderrortag__idtagparent",
                "iderror__iderrortag__tagtext",
                "iderror__iderrortag__tagtextrussian",
                "iderror__iderrortag__tagcolor",
                "iderror__iderrortag__idtagparent",
            )
            .filter(
                Q(iderror__iderrortag__isnull=False)
                & Q(iderror__iderrorlevel__isnull=False)
                & Q(iderror__iderrorlevel__errorlevelvalue=1)
                & Q(iderror__iderrorlevel__isnull=False)
            )
            .annotate(count_data=Count("iderror__iderrortag"))
        )

        # Пересчёт процента ошибок на токены

        data_on_tokens = dashboards.get_data_on_tokens(
            data_count_errors,
            id_data="iderrortag",
            is_unique_data=True,
            is_for_one_group=False,
        )

        # Построение иерархии ошибок
        data = dashboards.get_data_errors(data_on_tokens, level=0, is_sorted=True)

        # Получение словарей родителей и детей тегов
        tag_parents, dict_children = dashboards.get_dict_children()

        return render(
            request,
            "dashboard_error_types.html",
            {
                "right": True,
                "levels": levels,
                "groups": groups,
                "courses": courses,
                "texts": texts,
                "text_types": text_types,
                "data": data,
                "tag_parents": tag_parents,
                "dict_children": dict_children,
            },
        )

    else:
        list_filters = json.loads(request.body)
        flag_post = list_filters["flag_post"]

        if flag_post == "enrollment_date":
            enrollment_date = dashboards.get_enrollment_date(list_filters)
            return JsonResponse({"enrollment_date": enrollment_date}, status=200)

        if flag_post == "choice_all":
            texts, text_types = dashboards.get_filters_for_choice_all(list_filters)
            return JsonResponse({"texts": texts, "text_types": text_types}, status=200)

        if flag_post == "choice_group":
            texts, text_types = dashboards.get_filters_for_choice_group(list_filters)
            return JsonResponse({"texts": texts, "text_types": text_types}, status=200)

        if flag_post == "choice_student":
            texts, text_types = dashboards.get_filters_for_choice_student(list_filters)
            return JsonResponse({"texts": texts, "text_types": text_types}, status=200)

        if flag_post == "choice_course":
            texts, text_types = dashboards.get_filters_for_choice_course(list_filters)
            return JsonResponse({"texts": texts, "text_types": text_types}, status=200)

        if flag_post == "choice_text":
            groups, courses, text_types = dashboards.get_filters_for_choice_text(
                list_filters
            )
            return JsonResponse(
                {"groups": groups, "courses": courses, "text_types": text_types},
                status=200,
            )

        if flag_post == "choice_text_type":
            groups, courses, texts = dashboards.get_filters_for_choice_text_type(
                list_filters
            )
            return JsonResponse(
                {"groups": groups, "courses": courses, "texts": texts}, status=200
            )

        if flag_post == "update_diagrams":
            group = list_filters.get("group")
            date = list_filters.get("enrollment_date")
            surname = list_filters.get("surname")
            name = list_filters.get("name")
            course = list_filters.get("course")
            text = list_filters.get("text")
            text_type = list_filters.get("text_type")
            level = int(list_filters.get("level", 0))

            academic_year_title = None
            if date:
                start_year = date.split(" \\ ")[0]
                academic_year_title = f"{start_year}/{int(start_year) + 1}"

            base_filter = Q(iderrortag__markuptype=1) & Q(
                idsentence__idtext__errorcheckflag=True
            )

            if surname and name and text and text_type:
                base_filter &= (
                    Q(idsentence__idtext__idstudent__iduser__lastname=surname)
                    & Q(idsentence__idtext__idstudent__iduser__firstname=name)
                    & Q(idsentence__idtext__header=text)
                    & Q(idsentence__idtext__idtexttype__texttypename=text_type)
                )
            elif surname and name and text:
                base_filter &= (
                    Q(idsentence__idtext__idstudent__iduser__lastname=surname)
                    & Q(idsentence__idtext__idstudent__iduser__firstname=name)
                    & Q(idsentence__idtext__header=text)
                )
            elif surname and name and text_type:
                base_filter &= (
                    Q(idsentence__idtext__idstudent__iduser__lastname=surname)
                    & Q(idsentence__idtext__idstudent__iduser__firstname=name)
                    & Q(idsentence__idtext__idtexttype__texttypename=text_type)
                )
            elif surname and name:
                base_filter &= Q(
                    idsentence__idtext__idstudent__iduser__lastname=surname
                ) & Q(idsentence__idtext__idstudent__iduser__firstname=name)
            elif course and text and text_type:
                base_filter &= (
                    Q(idsentence__idtext__idstudent__idgroup__studycourse=course)
                    & Q(idsentence__idtext__header=text)
                    & Q(idsentence__idtext__idtexttype__texttypename=text_type)
                )
            elif course and text:
                base_filter &= Q(
                    idsentence__idtext__idstudent__idgroup__studycourse=course
                ) & Q(idsentence__idtext__header=text)
            elif course and text_type:
                base_filter &= Q(
                    idsentence__idtext__idstudent__idgroup__studycourse=course
                ) & Q(idsentence__idtext__idtexttype__texttypename=text_type)
            elif course:
                base_filter &= Q(
                    idsentence__idtext__idstudent__idgroup__studycourse=course
                )
            elif group and text and text_type and academic_year_title:
                base_filter &= (
                    Q(idsentence__idtext__idstudent__idgroup__groupname=group)
                    & Q(
                        idsentence__idtext__idstudent__idgroup__idayear__title=academic_year_title
                    )
                    & Q(idsentence__idtext__header=text)
                    & Q(idsentence__idtext__idtexttype__texttypename=text_type)
                )
            elif group and text and academic_year_title:
                base_filter &= (
                    Q(idsentence__idtext__idstudent__idgroup__groupname=group)
                    & Q(
                        idsentence__idtext__idstudent__idgroup__idayear__title=academic_year_title
                    )
                    & Q(idsentence__idtext__header=text)
                )
            elif group and text_type and academic_year_title:
                base_filter &= (
                    Q(idsentence__idtext__idstudent__idgroup__groupname=group)
                    & Q(
                        idsentence__idtext__idstudent__idgroup__idayear__title=academic_year_title
                    )
                    & Q(idsentence__idtext__idtexttype__texttypename=text_type)
                )
            elif group and academic_year_title:
                base_filter &= Q(
                    idsentence__idtext__idstudent__idgroup__groupname=group
                ) & Q(
                    idsentence__idtext__idstudent__idgroup__idayear__title=academic_year_title
                )
            elif text and text_type:
                base_filter &= Q(idsentence__idtext__header=text) & Q(
                    idsentence__idtext__idtexttype__texttypename=text_type
                )
            elif text_type:
                base_filter &= Q(idsentence__idtext__idtexttype__texttypename=text_type)
            elif text:
                base_filter &= Q(idsentence__idtext__header=text)

            data_count_errors = list(
                Sentence.objects.filter(base_filter)
                .values(
                    "iderrortag__iderrortag",
                    "iderrortag__idtagparent",
                    "iderrortag__tagtext",
                    "iderrortag__tagtextrussian",
                    "idtext",
                )
                .annotate(count_data=Count("iderrortag__iderrortag"))
            )

            data_on_tokens = dashboards.get_data_on_tokens(
                data_count_errors, "iderrortag__iderrortag", None, True, False
            )
            data = dashboards.get_data_errors(data_on_tokens, level, True)

            return JsonResponse({"data_type_errors": data}, status=200)


def chart_types_errors(request):	
	if request.method != 'POST':
		languages = ['Deustache']
		levels = dashboards.get_levels()
		groups = list(Group.objects.values('groupname').distinct().order_by('groupname'))
		courses = list(
			Group.objects.values('studycourse').filter(studycourse__gt=0).distinct().order_by(
				'studycourse'))
		texts = list(
			Text.objects.values('header').filter(errorcheckflag=True).distinct().order_by('header'))
		text_types = list(
			TextType.objects.values().filter(text__errorcheckflag=True).distinct().order_by('idtexttype'))
		
		data_count_errors = list(
			 Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
						  'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext')
						#   .filter(Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1))
						  .filter(errortoken__idtoken__idsentence__idtext__errorcheckflag=True)
						  .annotate(count_data=Count('iderrortag__iderrortag')))
		
		data_on_tokens = dashboards.get_data_on_tokens(data_count_errors, 'iderrortag__iderrortag', True,
							       False)
		data = dashboards.get_data_errors(data_on_tokens, 0, True)
		
		tag_parents, dict_children = dashboards.get_dict_children()
		
		return render(request, 'dashboard_error_types.html', {'right': True, 'levels': levels,
								      'groups': groups, 'courses': courses, 'texts': texts,
								      'text_types': text_types, 'data': data,
								      'tag_parents': tag_parents,
								      'dict_children': dict_children})
		
	else:
		list_filters = json.loads(request.body)
		flag_post = list_filters['flag_post']
		
		if flag_post == 'enrollment_date':
			enrollment_date = dashboards.get_enrollment_date(list_filters)
			return JsonResponse({'enrollment_date': enrollment_date}, status=200)
			
		if flag_post == 'choice_all':
			texts, text_types = dashboards.get_filters_for_choice_all(list_filters)
			return JsonResponse({'texts': texts, 'text_types': text_types}, status=200)
			
		if flag_post == 'choice_group':
			texts, text_types = dashboards.get_filters_for_choice_group(list_filters)
			return JsonResponse({'texts': texts, 'text_types': text_types}, status=200)
			
		if flag_post == 'choice_student':
			texts, text_types = dashboards.get_filters_for_choice_student(list_filters)
			return JsonResponse({'texts': texts, 'text_types': text_types}, status=200)
			
		if flag_post == 'choice_course':
			texts, text_types = dashboards.get_filters_for_choice_course(list_filters)
			return JsonResponse({'texts': texts, 'text_types': text_types}, status=200)
			
		if flag_post == 'choice_text':
			groups, courses, text_types = dashboards.get_filters_for_choice_text(list_filters)
			return JsonResponse({'groups': groups, 'courses': courses, 'text_types': text_types}, status=200)
			
		if flag_post == 'choice_text_type':
			groups, courses, texts = dashboards.get_filters_for_choice_text_type(list_filters)
			return JsonResponse({'groups': groups, 'courses': courses, 'texts': texts}, status=200)
			
		if flag_post == 'update_diagrams':
			group = list_filters['group']
			date = list_filters['enrollment_date']
			surname = list_filters['surname']
			name = list_filters['name']
			patronymic = list_filters['patronymic']
			course = list_filters['course']
			text = list_filters['text']
			text_type = list_filters['text_type']
			level = int(list_filters['level'])
			
			if surname and name and patronymic and text and text_type:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__user__last_name=surname) & Q(sentence__text_id__user__name=name) & Q(
							sentence__text_id__user__patronymic=patronymic) & Q(sentence__text_id__header=text) & Q(
							sentence__text_id__text_type=text_type)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif surname and name and patronymic and text:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__user__last_name=surname) & Q(sentence__text_id__user__name=name) & Q(
							sentence__text_id__user__patronymic=patronymic) & Q(
							sentence__text_id__header=text)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif surname and name and patronymic and text_type:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__user__last_name=surname) & Q(sentence__text_id__user__name=name) & Q(
							sentence__text_id__user__patronymic=patronymic) & Q(
							sentence__text_id__text_type=text_type)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif surname and name and patronymic:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__user__last_name=surname) & Q(sentence__text_id__user__name=name) & Q(
							sentence__text_id__user__patronymic=patronymic)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif surname and name and text and text_type:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__user__last_name=surname) & Q(sentence__text_id__user__name=name) & Q(
							sentence__text_id__header=text) & Q(sentence__text_id__text_type=text_type)).annotate(
						count_data=Count('iderrortag__iderrortag')))
				
			elif surname and name and text:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__user__last_name=surname) & Q(sentence__text_id__user__name=name) & Q(
							sentence__text_id__header=text)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif surname and name and text_type:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__user__last_name=surname) & Q(sentence__text_id__user__name=name) & Q(
							sentence__text_id__text_type=text_type)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif surname and name:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__user__last_name=surname) & Q(
							sentence__text_id__user__name=name)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif course and text_type and text:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__tbltextgroup__group__course_number=course) & Q(
							sentence__text_id__header=text) & Q(sentence__text_id__text_type=text_type)).annotate(
						count_data=Count('iderrortag__iderrortag')))
				
			elif course and text:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__tbltextgroup__group__course_number=course) & Q(
							sentence__text_id__header=text)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif course and text_type:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__tbltextgroup__group__course_number=course) & Q(
							sentence__text_id__text_type=text_type)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif course:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__tbltextgroup__group__course_number=course)).annotate(
						count_data=Count('iderrortag__iderrortag')))
				
			elif group and text and text_type:
				group_date = date[:4] + '-09-01'
				
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__tbltextgroup__group__group_name=group) & Q(
							sentence__text_id__tbltextgroup__group__enrollment_date=group_date) & Q(
							sentence__text_id__header=text) & Q(sentence__text_id__text_type=text_type)).annotate(
						count_data=Count('iderrortag__iderrortag')))
				
			elif group and text:
				group_date = date[:4] + '-09-01'
				
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__tbltextgroup__group__group_name=group) & Q(
							sentence__text_id__tbltextgroup__group__enrollment_date=group_date) & Q(
							sentence__text_id__header=text)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif group and text_type:
				group_date = date[:4] + '-09-01'
				
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__tbltextgroup__group__group_name=group) & Q(
							sentence__text_id__tbltextgroup__group__enrollment_date=group_date) & Q(
							sentence__text_id__text_type=text_type)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif group:
				group_date = date[:4] + '-09-01'
				
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__tbltextgroup__group__group_name=group) & Q(
							sentence__text_id__tbltextgroup__group__enrollment_date=group_date)).annotate(
						count_data=Count('iderrortag__iderrortag')))
				
			elif text_type and text:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__text_type=text_type) & Q(sentence__text_id__header=text)).annotate(
						count_data=Count('iderrortag__iderrortag')))
				
			elif text_type:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__text_type=text_type)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif text:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__header=text)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			else:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1)).annotate(
						count_data=Count('iderrortag__iderrortag')))
				
			data_on_tokens = dashboards.get_data_on_tokens(data_count_errors, 'iderrortag__iderrortag', True,
								       False)
			data = dashboards.get_data_errors(data_on_tokens, level, True)
			
			return JsonResponse({'data_type_errors': data}, status=200)


def chart_grade_errors(request):
	if request.method != 'POST':
		languages = ['Deustache']
		#groups = list(Group.objects.values('groupname', 'idyear__title').distinct().order_by('groupname'))
		groups = list(Group.objects.values('groupname').distinct().order_by('groupname'))
		courses = list(
			Group.objects.values('studycourse').filter(studycourse__gt=0).distinct().order_by(
				'studycourse'))
		texts = list(
			Text.objects.values('header').filter(errorcheckflag=True).distinct().order_by('header'))
		text_types = list(
			TextType.objects.values('idtexttype','texttypename').filter(text__errorcheckflag=True).distinct().order_by('idtexttype'))

		
		data_errorlevel = list(Error.objects.values('iderrorlevel__iderrorlevel', 'iderrorlevel__errorlevelname', 
							   'errortoken__idtoken__idsentence__idtext__idtext')
							#    .filter(Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(grade__id_grade__isnull=False))
							.filter(
								errortoken__idtoken__idsentence__idtext__errorcheckflag=True,  # проверенные тексты
								iderrorlevel__isnull=False
							)
							.annotate(count_data=Count('iderrorlevel__iderrorlevel')))
		
		data_errorlevel = dashboards.get_data_on_tokens(data_errorlevel, 'iderrorlevel__iderrorlevel',  True, False)
		data_errorlevel = dashboards.get_zero_count_grade_errors(data_errorlevel)
		data_errorlevel = sorted(data_errorlevel, key=lambda d: d['count_data'], reverse=True)


		return render(request, 'dashboard_error_grade.html', {'right': True, 'languages': languages, 'groups': groups,
								      'courses': courses, 'texts': texts,
								      'text_types': text_types, 'data': data_errorlevel})
	else:
		list_filters = json.loads(request.body)
		flag_post = list_filters['flag_post']
		
		if flag_post == 'enrollment_date':
			enrollment_date = dashboards.get_enrollment_date(list_filters)
			return JsonResponse({'enrollment_date': enrollment_date}, status=200)
			
		if flag_post == 'choice_all':
			texts, text_types = dashboards.get_filters_for_choice_all(list_filters)
			# print(texts, text_types)
			return JsonResponse({'texts': texts, 'text_types': text_types}, status=200)
			
		if flag_post == 'choice_group':
			texts, text_types = dashboards.get_filters_for_choice_group(list_filters)
			return JsonResponse({'texts': texts, 'text_types': text_types}, status=200)
			
		if flag_post == 'choice_student':
			texts, text_types = dashboards.get_filters_for_choice_student(list_filters)
			return JsonResponse({'texts': texts, 'text_types': text_types}, status=200)
			
		if flag_post == 'choice_course':
			texts, text_types = dashboards.get_filters_for_choice_course(list_filters)
			return JsonResponse({'texts': texts, 'text_types': text_types}, status=200)
			
		if flag_post == 'choice_text':
			groups, courses, text_types = dashboards.get_filters_for_choice_text(list_filters)
			return JsonResponse({'groups': groups, 'courses': courses, 'text_types': text_types}, status=200)
			
		if flag_post == 'choice_text_type':
			groups, courses, texts = dashboards.get_filters_for_choice_text_type(list_filters)
			return JsonResponse({'groups': groups, 'courses': courses, 'texts': texts}, status=200)
			
		
		if flag_post == 'update_diagrams':
			group = list_filters['group']
			date = list_filters['enrollment_date']
			surname = list_filters['surname']
			name = list_filters['name']
			patronymic = list_filters['patronymic']
			course = list_filters['course']
			text = list_filters['text']
			text_type = list_filters.get('text_type')
			
			
			if surname and name and patronymic and text and text_type:
				data_errorlevel = list(
					Error.objects.values('iderrorlevel__iderrorlevel', 'iderrorlevel__errorlevelname', 
								 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(errortoken__idtoken__idsentence__idtext__errorcheckflag=True) & Q(
							errortoken__idtoken__idsentence__idtext__idstudent__iduser__lastname=surname) & Q(errortoken__idtoken__idsentence__idtext__idstudent__iduser__firstname=name) & Q(
							errortoken__idtoken__idsentence__idtext__idstudent__iduser__middlename=patronymic) & Q(errortoken__idtoken__idsentence__idtext__header=text) & Q(
							errortoken__idtoken__idsentence__idtext__idtexttype__texttypename=text_type)).annotate(count_data=Count('iderrorlevel__iderrorlevel')))
				
			elif surname and name and patronymic and text: #
				data_errorlevel = list(
					Error.objects.values('iderrorlevel__iderrorlevel', 'iderrorlevel__errorlevelname', 
								 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(errortoken__idtoken__idsentence__idtext__errorcheckflag=True) & Q(
							errortoken__idtoken__idsentence__idtext__idstudent__iduser__lastname=surname) & Q(errortoken__idtoken__idsentence__idtext__idstudent__iduser__firstname=name) & Q(
							errortoken__idtoken__idsentence__idtext__idstudent__iduser__middlename=patronymic) & Q(
							errortoken__idtoken__idsentence__idtext__header=text)).annotate(count_data=Count('iderrorlevel__iderrorlevel')))
				
			elif surname and name and patronymic and text_type: 
				data_errorlevel = list(
					Error.objects.values('iderrorlevel__iderrorlevel', 'iderrorlevel__errorlevelname', 
								 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(errortoken__idtoken__idsentence__idtext__errorcheckflag=True) & Q(
							errortoken__idtoken__idsentence__idtext__idstudent__iduser__lastname=surname) & Q(errortoken__idtoken__idsentence__idtext__idstudent__iduser__firstname=name) & Q(
							errortoken__idtoken__idsentence__idtext__idstudent__iduser__middlename=patronymic) & Q(
							errortoken__idtoken__idsentence__idtext__idtexttype__texttypename=text_type)).annotate(count_data=Count('iderrorlevel__iderrorlevel')))
				
			elif surname and name and patronymic:
				data_errorlevel = list(
					Error.objects.values('iderrorlevel__iderrorlevel', 'iderrorlevel__errorlevelname', 
								 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(errortoken__idtoken__idsentence__idtext__errorcheckflag=True) & Q(
							errortoken__idtoken__idsentence__idtext__idstudent__iduser__lastname=surname) & Q(errortoken__idtoken__idsentence__idtext__idstudent__iduser__firstname=name) & Q(
							errortoken__idtoken__idsentence__idtext__idstudent__iduser__middlename=patronymic)).annotate(
						count_data=Count('iderrorlevel__iderrorlevel')))
				
			elif surname and name and text and text_type:
				data_errorlevel = list(
					Error.objects.values('iderrorlevel__iderrorlevel', 'iderrorlevel__errorlevelname', 
								 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(errortoken__idtoken__idsentence__idtext__errorcheckflag=True) & Q(
							errortoken__idtoken__idsentence__idtext__idstudent__iduser__lastname=surname) & Q(errortoken__idtoken__idsentence__idtext__idstudent__iduser__firstname=name) & Q(
							errortoken__idtoken__idsentence__idtext__header=text) & Q(errortoken__idtoken__idsentence__idtext__idtexttype__texttypename=text_type)).annotate(
						count_data=Count('iderrorlevel__iderrorlevel')))
				
			elif surname and name and text:
				data_errorlevel = list(
					Error.objects.values('iderrorlevel__iderrorlevel', 'iderrorlevel__errorlevelname', 
								 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						 Q(errortoken__idtoken__idsentence__idtext__errorcheckflag=True) & Q(
							errortoken__idtoken__idsentence__idtext__idstudent__iduser__lastname=surname) & Q(errortoken__idtoken__idsentence__idtext__idstudent__iduser__firstname=name) & Q(
							errortoken__idtoken__idsentence__idtext__header=text)).annotate(count_data=Count('iderrorlevel__iderrorlevel')))
				
			elif surname and name and text_type:
				data_errorlevel = list(
					Error.objects.values('iderrorlevel__iderrorlevel', 'iderrorlevel__errorlevelname', 
								 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(errortoken__idtoken__idsentence__idtext__errorcheckflag=True) &  Q(
							errortoken__idtoken__idsentence__idtext__idstudent__iduser__lastname=surname) & Q(
							errortoken__idtoken__idsentence__idtext__idstudent__iduser__firstname=name) & Q(
							errortoken__idtoken__idsentence__idtext__idtexttype__texttypename=text_type)).annotate(count_data=Count('iderrorlevel__iderrorlevel')))
				
			elif surname and name:
				data_errorlevel = list(
					Error.objects.values('iderrorlevel__iderrorlevel', 'iderrorlevel__errorlevelname',
						   'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(errortoken__idtoken__idsentence__idtext__errorcheckflag=True) & Q(
							errortoken__idtoken__idsentence__idtext__idstudent__iduser__lastname=surname) & Q(
							errortoken__idtoken__idsentence__idtext__idstudent__iduser__firstname=name)).annotate(count_data=Count('iderrorlevel__iderrorlevel')))
				
			elif course and text_type and text:
				data_errorlevel = list(
					Error.objects.values('iderrorlevel__iderrorlevel', 'iderrorlevel__errorlevelname', 
								 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(errortoken__idtoken__idsentence__idtext__errorcheckflag=True) & Q(
							errortoken__idtoken__idsentence__idtext__idstudent__idgroup__studycourse=course) & Q(
							errortoken__idtoken__idsentence__idtext__header=text) & Q(errortoken__idtoken__idsentence__idtext__idtexttype__texttypename=text_type)).annotate(
						count_data=Count('iderrorlevel__iderrorlevel')))
				
			elif course and text:
				data_errorlevel = list(
					Error.objects.values('iderrorlevel__iderrorlevel', 'iderrorlevel__errorlevelname',
						   'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(errortoken__idtoken__idsentence__idtext__errorcheckflag=True) & Q(
							errortoken__idtoken__idsentence__idtext__idstudent__idgroup__studycourse=course) & Q(
							errortoken__idtoken__idsentence__idtext__header=text)).annotate(count_data=Count('iderrorlevel__iderrorlevel')))
				
			elif course and text_type:
				data_errorlevel = list(
					Error.objects.values('iderrorlevel__iderrorlevel', 'iderrorlevel__errorlevelname',
						   'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(errortoken__idtoken__idsentence__idtext__errorcheckflag=True) & Q(
							errortoken__idtoken__idsentence__idtext__idstudent__idgroup__studycourse=course) & Q(
							errortoken__idtoken__idsentence__idtext__idtexttype__texttypename=text_type)).annotate(count_data=Count('iderrorlevel__iderrorlevel')))
				
			elif course:
				data_errorlevel = list(
					Error.objects.values( 'iderrorlevel__iderrorlevel', 'iderrorlevel__errorlevelname',
						   'errortoken__idtoken__idsentence__idtext__idtext').filter(Q(errortoken__idtoken__idsentence__idtext__errorcheckflag=True) & 
    Q(errortoken__idtoken__idsentence__idtext__idstudent__idgroup__studycourse=course)).annotate(count_data=Count('iderrorlevel__iderrorlevel')))

				
			elif group and text and text_type: #!!!!
				group_date = date[:4] + '-09-01'
				data_errorlevel = list(
					Error.objects.values('iderrorlevel__iderrorlevel', 'iderrorlevel__errorlevelname', 
								 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						 Q(errortoken__idtoken__idsentence__idtext__errorcheckflag=True) & Q(
							errortoken__idtoken__idsentence__idtext__idstudent__idgroup__groupname=group) & Q(
							errortoken__idtoken__idsentence__idtext__idstudent__idgroup__idayear__title=group_date) & Q(
							sentence__text_id__header=text) & Q(errortoken__idtoken__idsentence__idtext__idtexttype__texttypename=text_type)).annotate(
						count_data=Count('iderrorlevel__iderrorlevel')))
				
			elif group and text: # !!!!!
				group_date = date[0:4]+"/"+date[7:]
				data_errorlevel = list(
					Error.objects.values('iderrorlevel__iderrorlevel', 'iderrorlevel__errorlevelname', 
								 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						 Q(errortoken__idtoken__idsentence__idtext__errorcheckflag=True) & Q(
							errortoken__idtoken__idsentence__idtext__idstudent__idgroup__groupname=group) & Q(
							errortoken__idtoken__idsentence__idtext__idstudent__idgroup__idayear__title=group_date) & Q(
							errortoken__idtoken__idsentence__idtext__header=text)).annotate(count_data=Count('iderrorlevel__iderrorlevel')))
				
			elif group and text_type: # !!!!!!!
				group_date = date[0:4]+"/"+date[7:]
				data_errorlevel = list(
					Error.objects.values('iderrorlevel__iderrorlevel', 'iderrorlevel__errorlevelname', 
								 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						 Q(errortoken__idtoken__idsentence__idtext__errorcheckflag=True) & Q(
							errortoken__idtoken__idsentence__idtext__idstudent__idgroup__groupname=group) & Q(
							errortoken__idtoken__idsentence__idtext__idstudent__idgroup__idayear__title=group_date) & Q(
							errortoken__idtoken__idsentence__idtext__idtexttype__texttypename=text_type)).annotate(count_data=Count('iderrorlevel__iderrorlevel')))
				
			elif group: #
				group_date = date
				group_date = date[0:4]+"/"+date[7:]
				data_errorlevel = list(
					Error.objects.values('iderrorlevel__iderrorlevel', 'iderrorlevel__errorlevelname', 
								 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						 Q(errortoken__idtoken__idsentence__idtext__errorcheckflag=True) & Q(
							errortoken__idtoken__idsentence__idtext__idstudent__idgroup__groupname=group) & Q(
							errortoken__idtoken__idsentence__idtext__idstudent__idgroup__idayear__title=group_date)).annotate(
						count_data=Count('iderrorlevel__iderrorlevel')))
				
			elif text_type and text: 
				data_errorlevel = list(
					Error.objects.values('iderrorlevel__iderrorlevel', 'iderrorlevel__errorlevelname', 
								 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						 Q(errortoken__idtoken__idsentence__idtext__errorcheckflag=True) & Q(
							errortoken__idtoken__idsentence__idtext__idtexttype__texttypename=text_type) & Q(errortoken__idtoken__idsentence__idtext__header=text)).annotate(
						count_data=Count('iderrorlevel__iderrorlevel')))
				
			elif text_type: 
				data_errorlevel = list(
					Error.objects.values('iderrorlevel__iderrorlevel', 'iderrorlevel__errorlevelname', 
								 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						 Q(errortoken__idtoken__idsentence__idtext__errorcheckflag=True) & Q(
							errortoken__idtoken__idsentence__idtext__idtexttype__texttypename=text_type)).annotate(count_data=Count('iderrorlevel__iderrorlevel')))
				
			elif text: 
				data_errorlevel = list(
					Error.objects.values('iderrorlevel__iderrorlevel', 'iderrorlevel__errorlevelname', 
								 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						 Q(errortoken__idtoken__idsentence__idtext__errorcheckflag=True) & Q(
							errortoken__idtoken__idsentence__idtext__header=text)).annotate(count_data=Count('iderrorlevel__iderrorlevel')))
				
			else: 
				data_errorlevel = list(
					Error.objects.values('iderrorlevel__iderrorlevel', 'iderrorlevel__errorlevelname', 
								 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						 Q(errortoken__idtoken__idsentence__idtext__errorcheckflag=True)).annotate(
						count_data=Count('iderrorlevel__iderrorlevel')))
				
			data_errorlevel = dashboards.get_data_on_tokens(data_errorlevel, 'iderrorlevel__iderrorlevel',  True,
								   False)
			data_errorlevel = dashboards.get_zero_count_grade_errors(data_errorlevel)
			data_errorlevel = sorted(data_errorlevel, key=lambda d: d['count_data'], reverse=True)
			print(data_errorlevel, "!!!!")
			return JsonResponse({'data_grade_errors': data_errorlevel}, status=200)


def chart_types_grade_errors(request):		
	if request.method != 'POST':
		languages = ['Deustache']
		levels = dashboards.get_levels()
		groups = list(Group.objects.values('groupname').distinct().order_by('groupname'))
		courses = list(
			Group.objects.values('studycourse').filter(studycourse__gt=0).distinct().order_by(
				'studycourse'))
		texts = list(
			Text.objects.values('header').filter(errorcheckflag=True).distinct().order_by('header'))
		text_types = list(
			TextType.objects.values().filter(text__errorcheckflag=True).distinct().order_by('idtexttype'))
		grades = list(ErrorLevel.objects.values('iderrorlevel', 'errorlevelname').order_by('errorlevelname'))
		
		data_on_tokens = []
		texts_id = {}
		# count_grades = {}
		count_grades = len(grades)
		for grade in grades:
			data_count_errors = list(
				Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
							 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext')
							#  .filter(Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							# 		grade=grade["id_grade"])).annotate(count_data=Count('iderrortag__iderrortag')))
							.filter(
									errortoken__idtoken__idsentence__idtext__errorcheckflag=True,
									iderrorlevel=grade["iderrorlevel"]
								)
							.annotate(count_data=Count('iderrortag__iderrortag')))
			
			texts_id = dashboards.get_texts_id_keys(data_count_errors, texts_id)
			data_count_on_tokens, texts_id = dashboards.get_texts_id_and_data_on_tokens(data_count_errors, texts_id,
												    'iderrortag__iderrortag')
			data_on_tokens.append(data_count_on_tokens)
			
			# if grade['grade_language'] not in count_grades_for_language.keys():
			# 	count_grades_for_language[grade['grade_language']] = 1
			# else:
			# 	count_grades_for_language[grade['grade_language']] += 1
			# if 'grade_count' not in count_grades:
			# 	count_grades['grade_count'] = 1
			# else:
			# 	count_grades['grade_count'] += 1
		
		data = []
		for i in range(len(data_on_tokens)):
			data_count = dashboards.get_on_tokens(texts_id, data_on_tokens[i])
			data.append(dashboards.get_data_errors(data_count, 0, False))
			
		for i in range(len(data[0])):
			sum_count = sum(item[i]['count_data'] for item in data)
			for item in data:
				item[i]['sum_count'] = sum_count
			# offset = 0
			# for language in count_grades_for_language.keys():
			# 	sum_count = 0
			# 	for j in range(count_grades_for_language[language]):
			# 		sum_count += data[offset + j][i]['count_data']
					
			# 	for j in range(count_grades_for_language[language]):
			# 		data[offset + j][i]['sum_count'] = sum_count
					
			# 	offset += count_grades_for_language[language]
				
		for i in range(len(data)):
			data[i] = sorted(data[i], key=lambda d: d['sum_count'], reverse=True)
			
		tag_parents, dict_children = dashboards.get_dict_children()
		
		return render(request, 'dashboard_error_types_grade.html', {'right': True, 'languages': languages,
									    'levels': levels, 'groups': groups,
									    'courses': courses, 'texts': texts,
									    'text_types': text_types, 'data': data,
									    'grades': grades, 'tag_parents': tag_parents,
									    'dict_children': dict_children,
									    'count_grades_for_language':
									    	count_grades})
	else:
		list_filters = json.loads(request.body)
		flag_post = list_filters['flag_post']
		
		if flag_post == 'enrollment_date':
			enrollment_date = dashboards.get_enrollment_date(list_filters)
			return JsonResponse({'enrollment_date': enrollment_date}, status=200)
			
		if flag_post == 'choice_all':
			texts, text_types = dashboards.get_filters_for_choice_all(list_filters)
			return JsonResponse({'texts': texts, 'text_types': text_types}, status=200)
			
		if flag_post == 'choice_group':
			texts, text_types = dashboards.get_filters_for_choice_group(list_filters)
			return JsonResponse({'texts': texts, 'text_types': text_types}, status=200)
			
		if flag_post == 'choice_student':
			texts, text_types = dashboards.get_filters_for_choice_student(list_filters)
			return JsonResponse({'texts': texts, 'text_types': text_types}, status=200)
			
		if flag_post == 'choice_course':
			texts, text_types = dashboards.get_filters_for_choice_course(list_filters)
			return JsonResponse({'texts': texts, 'text_types': text_types}, status=200)
			
		if flag_post == 'choice_text':
			groups, courses, text_types = dashboards.get_filters_for_choice_text(list_filters)
			return JsonResponse({'groups': groups, 'courses': courses, 'text_types': text_types}, status=200)
			
		if flag_post == 'choice_text_type':
			groups, courses, texts = dashboards.get_filters_for_choice_text_type(list_filters)
			return JsonResponse({'groups': groups, 'courses': courses, 'texts': texts}, status=200)
			
		if flag_post == 'update_diagrams':
			group = list_filters['group']
			date = list_filters['enrollment_date']
			surname = list_filters['surname']
			name = list_filters['name']
			patronymic = list_filters['patronymic']
			course = list_filters['course']
			text = list_filters['text']
			text_type = list_filters['text_type']
			level = int(list_filters['level'])
			
			grades = list(
				ErrorLevel.objects.values('iderrorlevel', 'errorlevelname').order_by('errorlevelname'))
			
			data_on_tokens = []
			texts_id = {}
			count_grades = 0  # Просто счетчик вместо count_grades_for_language
			for grade in grades:
				base_query = Error.objects.values(
					'iderrortag__iderrortag', 
					'iderrortag__idtagparent', 
					'iderrortag__tagtext',
					'iderrortag__tagtextrussian', 
					'errortoken__idtoken__idsentence__idtext__idtext'
				).filter(
					errortoken__idtoken__idsentence__idtext__errorcheckflag=True,
					iderrorlevel=grade["iderrorlevel"]
				)
				
				# Упрощенная фильтрация
				if surname and name:
					base_query = base_query.filter(
						errortoken__idtoken__idsentence__idtext__idstudent__lastname=surname,
						errortoken__idtoken__idsentence__idtext__idstudent__firstname=name
					)
					if patronymic:
						base_query = base_query.filter(
							errortoken__idtoken__idsentence__idtext__idstudent__patronymic=patronymic
						)
				
				if course:
					base_query = base_query.filter(
						errortoken__idtoken__idsentence__idtext__idstudent__idgroup__studycourse=course
					)
				
				if group and date:
					group_date = date[:4] + '-09-01'
					base_query = base_query.filter(
						errortoken__idtoken__idsentence__idtext__idstudent__idgroup__groupname=group,
						errortoken__idtoken__idsentence__idtext__idstudent__idgroup__idayear__title=group_date
					)
				
				if text:
					base_query = base_query.filter(
						errortoken__idtoken__idsentence__idtext__header=text
					)
				
				if text_type:
					base_query = base_query.filter(
						errortoken__idtoken__idsentence__idtext__idtexttype=text_type
					)
				
				data_count_errors = list(base_query.annotate(
					count_data=Count('iderrortag__iderrortag')
				))
				
				texts_id = dashboards.get_texts_id_keys(data_count_errors, texts_id)
				data_count_on_tokens, texts_id = dashboards.get_texts_id_and_data_on_tokens(
					data_count_errors, texts_id, 'iderrortag__iderrortag'
				)
				data_on_tokens.append(data_count_on_tokens)
				count_grades += 1
			
			# Обработка данных
			data = []
			for i in range(len(data_on_tokens)):
				data_count = dashboards.get_on_tokens(texts_id, data_on_tokens[i])
				data.append(dashboards.get_data_errors(data_count, level, False))
			
			# Суммирование данных (упрощенная версия без разделения по языкам)
			for i in range(len(data[0])):
				sum_count = sum(item[i]['count_data'] for item in data)
				for item in data:
					item[i]['sum_count'] = sum_count
			
			# Сортировка
			for i in range(len(data)):
				data[i] = sorted(data[i], key=lambda d: d['sum_count'], reverse=True)
			
			return JsonResponse({'data': data}, status=200)
			# count_grades_for_language = {}
			# for grade in grades:
			# 	if surname and name and patronymic and text and text_type:
			# 		data_count_errors = list(
			# 			Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
			# 						 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
			# 				Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
			# 					grade=grade["id_grade"]) & Q(sentence__text_id__user__last_name=surname) & Q(
			# 					sentence__text_id__user__name=name) & Q(
			# 					sentence__text_id__user__patronymic=patronymic) & Q(sentence__text_id__header=text) & Q(
			# 					sentence__text_id__text_type=text_type)).annotate(count_data=Count('iderrortag__iderrortag')))
					
			# 	elif surname and name and patronymic and text:
			# 		data_count_errors = list(
			# 			Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
			# 						 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
			# 				Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
			# 					grade=grade["id_grade"]) & Q(sentence__text_id__user__last_name=surname) & Q(
			# 					sentence__text_id__user__name=name) & Q(
			# 					sentence__text_id__user__patronymic=patronymic) & Q(
			# 					sentence__text_id__header=text)).annotate(count_data=Count('iderrortag__iderrortag')))
					
			# 	elif surname and name and patronymic and text_type:
			# 		data_count_errors = list(
			# 			Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
			# 						 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
			# 				Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
			# 					grade=grade["id_grade"]) & Q(sentence__text_id__user__last_name=surname) & Q(
			# 					sentence__text_id__user__name=name) & Q(
			# 					sentence__text_id__user__patronymic=patronymic) & Q(
			# 					sentence__text_id__text_type=text_type)).annotate(count_data=Count('iderrortag__iderrortag')))
					
			# 	elif surname and name and patronymic:
			# 		data_count_errors = list(
			# 			Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
			# 						 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
			# 				Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
			# 					grade=grade["id_grade"]) & Q(sentence__text_id__user__last_name=surname) & Q(
			# 					sentence__text_id__user__name=name) & Q(
			# 					sentence__text_id__user__patronymic=patronymic)).annotate(
			# 				count_data=Count('iderrortag__iderrortag')))
					
			# 	elif surname and name and text and text_type:
			# 		data_count_errors = list(
			# 			Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
			# 						 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
			# 				Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
			# 					grade=grade["id_grade"]) & Q(sentence__text_id__user__last_name=surname) & Q(
			# 					sentence__text_id__user__name=name) & Q(sentence__text_id__header=text) & Q(
			# 					sentence__text_id__text_type=text_type)).annotate(count_data=Count('iderrortag__iderrortag')))
					
			# 	elif surname and name and text:
			# 		data_count_errors = list(
			# 			Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
			# 						 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
			# 				Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
			# 					grade=grade["id_grade"]) & Q(sentence__text_id__user__last_name=surname) & Q(
			# 					sentence__text_id__user__name=name) & Q(sentence__text_id__header=text)).annotate(
			# 				count_data=Count('iderrortag__iderrortag')))
					
			# 	elif surname and name and text_type:
			# 		data_count_errors = list(
			# 			Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
			# 						 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
			# 				Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
			# 					grade=grade["id_grade"]) & Q(sentence__text_id__user__last_name=surname) & Q(
			# 					sentence__text_id__user__name=name) & Q(
			# 					sentence__text_id__text_type=text_type)).annotate(count_data=Count('iderrortag__iderrortag')))
					
			# 	elif surname and name:
			# 		data_count_errors = list(
			# 			Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
			# 						 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
			# 				Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
			# 					grade=grade["id_grade"]) & Q(sentence__text_id__user__last_name=surname) & Q(
			# 					sentence__text_id__user__name=name)).annotate(count_data=Count('iderrortag__iderrortag')))
					
			# 	elif course and text_type and text:
			# 		data_count_errors = list(
			# 			Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
			# 						 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
			# 				Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
			# 					grade=grade["id_grade"]) & Q(
			# 					sentence__text_id__tbltextgroup__group__course_number=course) & Q(
			# 					sentence__text_id__header=text) & Q(sentence__text_id__text_type=text_type)).annotate(
			# 				count_data=Count('iderrortag__iderrortag')))
					
			# 	elif course and text:
			# 		data_count_errors = list(
			# 			Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
			# 						 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
			# 				Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
			# 					grade=grade["id_grade"]) & Q(
			# 					sentence__text_id__tbltextgroup__group__course_number=course) & Q(
			# 					sentence__text_id__header=text)).annotate(count_data=Count('iderrortag__iderrortag')))
					
			# 	elif course and text_type:
			# 		data_count_errors = list(
			# 			Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
			# 						 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
			# 				Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
			# 					grade=grade["id_grade"]) & Q(
			# 					sentence__text_id__tbltextgroup__group__course_number=course) & Q(
			# 					sentence__text_id__text_type=text_type)).annotate(count_data=Count('iderrortag__iderrortag')))
					
			# 	elif course:
			# 		data_count_errors = list(
			# 			Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
			# 						 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
			# 				Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
			# 					grade=grade["id_grade"]) & Q(
			# 					sentence__text_id__tbltextgroup__group__course_number=course)).annotate(
			# 				count_data=Count('iderrortag__iderrortag')))
					
			# 	elif group and text and text_type:
			# 		group_date = date[:4] + '-09-01'
			# 		data_count_errors = list(
			# 			Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
			# 						 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
			# 				Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
			# 					grade=grade["id_grade"]) & Q(
			# 					sentence__text_id__tbltextgroup__group__group_name=group) & Q(
			# 					sentence__text_id__tbltextgroup__group__enrollment_date=group_date) & Q(
			# 					sentence__text_id__header=text) & Q(sentence__text_id__text_type=text_type)).annotate(
			# 				count_data=Count('iderrortag__iderrortag')))
					
			# 	elif group and text:
			# 		group_date = date[:4] + '-09-01'
			# 		data_count_errors = list(
			# 			Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
			# 						 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
			# 				Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
			# 					grade=grade["id_grade"]) & Q(
			# 					sentence__text_id__tbltextgroup__group__group_name=group) & Q(
			# 					sentence__text_id__tbltextgroup__group__enrollment_date=group_date) & Q(
			# 					sentence__text_id__header=text)).annotate(count_data=Count('iderrortag__iderrortag')))
					
			# 	elif group and text_type:
			# 		group_date = date[:4] + '-09-01'
			# 		data_count_errors = list(
			# 			Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
			# 						 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
			# 				Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
			# 					grade=grade["id_grade"]) & Q(
			# 					sentence__text_id__tbltextgroup__group__group_name=group) & Q(
			# 					sentence__text_id__tbltextgroup__group__enrollment_date=group_date) & Q(
			# 					sentence__text_id__text_type=text_type)).annotate(count_data=Count('iderrortag__iderrortag')))
					
			# 	elif group:
			# 		group_date = date[:4] + '-09-01'
			# 		data_count_errors = list(
			# 			Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
			# 						 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
			# 				Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
			# 					grade=grade["id_grade"]) & Q(
			# 					sentence__text_id__tbltextgroup__group__group_name=group) & Q(
			# 					sentence__text_id__tbltextgroup__group__enrollment_date=group_date)).annotate(
			# 				count_data=Count('iderrortag__iderrortag')))
					
			# 	elif text_type and text:
			# 		data_count_errors = list(
			# 			Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
			# 						 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
			# 				Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
			# 					grade=grade["id_grade"]) & Q(sentence__text_id__text_type=text_type) & Q(
			# 					sentence__text_id__header=text)).annotate(count_data=Count('iderrortag__iderrortag')))
					
			# 	elif text_type:
			# 		data_count_errors = list(
			# 			Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
			# 						 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
			# 				Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
			# 					grade=grade["id_grade"]) & Q(sentence__text_id__text_type=text_type)).annotate(
			# 				count_data=Count('iderrortag__iderrortag')))
					
			# 	elif text:
			# 		data_count_errors = list(
			# 			Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
			# 						 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
			# 				Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
			# 					grade=grade["id_grade"]) & Q(sentence__text_id__header=text)).annotate(
			# 				count_data=Count('iderrortag__iderrortag')))
					
			# 	else:
			# 		data_count_errors = list(
			# 			Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
			# 						 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
			# 				Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
			# 					grade=grade["id_grade"])).annotate(count_data=Count('iderrortag__iderrortag')))
					
				# texts_id = dashboards.get_texts_id_keys(data_count_errors, texts_id)
				# data_count_on_tokens, texts_id = dashboards.get_texts_id_and_data_on_tokens(data_count_errors, texts_id,
				# 									    'iderrortag__iderrortag')
				# data_on_tokens.append(data_count_on_tokens)
				
				# if grade['grade_language'] not in count_grades_for_language.keys():
				# 	count_grades_for_language[grade['grade_language']] = 1
				# else:
				# 	count_grades_for_language[grade['grade_language']] += 1
			# data = []
			# for i in range(len(data_on_tokens)):
			# 	data_count = dashboards.get_on_tokens(texts_id, data_on_tokens[i])
			# 	data.append(dashboards.get_data_errors(data_count, level, False))
				
			# for i in range(len(data[0])):
			# 	offset = 0
			# 	for language in count_grades_for_language.keys():
			# 		sum_count = 0
			# 		for j in range(count_grades_for_language[language]):
			# 			sum_count += data[offset + j][i]['count_data']
						
			# 		for j in range(count_grades_for_language[language]):
			# 			data[offset + j][i]['sum_count'] = sum_count
						
			# 		offset += count_grades_for_language[language]
					
			# for i in range(len(data)):
			# 	data[i] = sorted(data[i], key=lambda d: d['sum_count'], reverse=True)
				
			# return JsonResponse({'data': data}, status=200)


# НУЖНО ПЕРЕПИСАТЬ ПОСЛЕ ELSE ЗАПРОСЫ
def chart_student_dynamics(request):		
	if request.method != 'POST':
		languages = ['Deustache']
		tags = list(Error.objects.values('iderrortag__iderrortag', 'iderrortag__tagtext', 'iderrortag__tagtextrussian').order_by('iderrortag__iderrortag').distinct())

		print(tags)
		
		return render(request, 'dashboard_student_dynamics.html', {'right': True, 'languages': languages, 'tags': tags})
	else:
		list_filters = json.loads(request.body)
		print(list_filters, "&&")
		text_type = list_filters['text_type']
		surname = list_filters['surname']
		name = list_filters['name']
		patronymic = list_filters['patronymic']
		tag = list_filters['tag']
		checked_tag_children = list_filters['checked_tag_children']
		
		tags = [tag]
		if checked_tag_children:
			tags = dashboards.get_tag_children(tag)
			
		list_text_id_with_markup = []
		if surname and name and patronymic and tag and text_type:
			data_count_errors = list(Error.objects.values('errortoken__idtoken__idsentence__idtext__createdate',
									  'errortoken__idtoken__idsentence__idtext__idtext')
									  .filter(
										errortoken__idtoken__idsentence__idtext__idstudent__lastname=surname,
										errortoken__idtoken__idsentence__idtext__idstudent__firstname=name,
										errortoken__idtoken__idsentence__idtext__idstudent__patronymic=patronymic,
										errortoken__idtoken__idsentence__idtext__idtexttype=text_type,
										iderrortag__in=tags,
        								errortoken__idtoken__idsentence__idtext__errorcheckflag=True
				# Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
				# 	sentence__text_id__user__last_name=surname) & Q(sentence__text_id__user__name=name) & Q(
				# 	sentence__text_id__user__patronymic=patronymic) & Q(tag__id_tag__in=tags) & Q(
				# 	sentence__text_id__text_type=text_type)
									)
					.annotate(count_data=Count('errortoken__idtoken__idsentence__idtext__createdate')))
			
			for data in data_count_errors:
				list_text_id_with_markup.append(data["sentence__text_id"])
				
			texts_without_markup = list(Text.objects.annotate(tag__tag_language=F('language'),
									     sentence__text_id__create_date=F('create_date'),
									     sentence__text_id=F('id_text')).values(
				'sentence__text_id__create_date', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
				Q(errorcheckflag=True) & Q(user__last_name=surname) & Q(user__name=name) & Q(
					user__patronymic=patronymic) & Q(text_type=text_type) & ~Q(
					id_text__in=list_text_id_with_markup)).annotate(count_data=Value(0, output_field=IntegerField())))
			
		elif surname and name and patronymic and tag:
			data_count_errors = list(Error.objects.values('sentence__text_id__create_date',
									  'errortoken__idtoken__idsentence__idtext__idtext').filter(
				Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
					sentence__text_id__user__last_name=surname) & Q(sentence__text_id__user__name=name) & Q(
					sentence__text_id__user__patronymic=patronymic) & Q(tag__id_tag__in=tags)).annotate(
				count_data=Count('sentence__text_id__create_date')))
			
			for data in data_count_errors:
				list_text_id_with_markup.append(data["sentence__text_id"])
				
			texts_without_markup = list(Text.objects.annotate(tag__tag_language=F('language'),
									     sentence__text_id__create_date=F('create_date'),
									     sentence__text_id=F('id_text')).values(
				'sentence__text_id__create_date', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
				Q(errorcheckflag=True) & Q(user__last_name=surname) & Q(user__name=name) & Q(
					user__patronymic=patronymic) & ~Q(id_text__in=list_text_id_with_markup)).annotate(
				count_data=Value(0, output_field=IntegerField())))
			
		elif surname and name and tag and text_type:
			data_count_errors = list(Error.objects.values('sentence__text_id__create_date',
									  'errortoken__idtoken__idsentence__idtext__idtext').filter(
				Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
					sentence__text_id__user__last_name=surname) & Q(sentence__text_id__user__name=name) & Q(
					tag__id_tag__in=tags) & Q(sentence__text_id__text_type=text_type)).annotate(
				count_data=Count('sentence__text_id__create_date')))
			
			for data in data_count_errors:
				list_text_id_with_markup.append(data["sentence__text_id"])
				
			texts_without_markup = list(Text.objects.annotate(tag__tag_language=F('language'),
									     sentence__text_id__create_date=F('create_date'),
									     sentence__text_id=F('id_text')).values(
				'sentence__text_id__create_date', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
				Q(errorcheckflag=True) & Q(user__last_name=surname) & Q(user__name=name) & Q(text_type=text_type) & ~Q(
					id_text__in=list_text_id_with_markup)).annotate(count_data=Value(0, output_field=IntegerField())))
			
		elif surname and name and tag:
			data_count_errors = list(Error.objects.values('sentence__text_id__create_date',
									  'errortoken__idtoken__idsentence__idtext__idtext').filter(
				Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
					sentence__text_id__user__last_name=surname) & Q(sentence__text_id__user__name=name) & Q(
					tag__id_tag__in=tags)).annotate(count_data=Count('sentence__text_id__create_date')))
			
			for data in data_count_errors:
				list_text_id_with_markup.append(data["sentence__text_id"])
				
			texts_without_markup = list(Text.objects.annotate(tag__tag_language=F('language'),
									     sentence__text_id__create_date=F('create_date'),
									     sentence__text_id=F('id_text')).values(
				'sentence__text_id__create_date', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
				Q(errorcheckflag=True) & Q(user__last_name=surname) & Q(user__name=name) & ~Q(
					id_text__in=list_text_id_with_markup)).annotate(count_data=Value(0, output_field=IntegerField())))
			
		data_count_errors.extend(texts_without_markup)
		data_count_errors = dashboards.get_data_on_tokens(data_count_errors, '', False, False)
		
		texts_with_create_date = []
		for data in data_count_errors:
			if data['sentence__text_id__create_date'] is None:
				texts_with_create_date.append(data)
				
		for text in texts_with_create_date:
			data_count_errors.remove(text)
			
		data_count_errors = sorted(data_count_errors, key=lambda d: d['sentence__text_id__create_date'])
		
		if patronymic:
			text_types = list(TextType.objects.values().filter(Q(text__errorcheckflag=True) & Q(
				tbltext__user__last_name=surname) & Q(tbltext__user__name=name) & Q(
				tbltext__user__patronymic=patronymic) & Q(
				tbltext__tblsentence__tblmarkup__tag__in=tags)).distinct().order_by('idtexttype'))
		else:
			text_types = list(TextType.objects.values().filter(Q(text__errorcheckflag=True) & Q(
				tbltext__user__last_name=surname) & Q(tbltext__user__name=name) & Q(
				tbltext__tblsentence__tblmarkup__tag__in=tags)).distinct().order_by('idtexttype'))
			
		return JsonResponse({'data': data_count_errors, 'text_types': text_types}, status=200)


def chart_groups_errors(request):		
	if request.method != 'POST':
		languages = ['Deustache']
		tags = list(Error.objects.values('iderrortag__iderrortag', 'iderrortag__tagtext', 'iderrortag__tagtextrussian').order_by('iderrortag__iderrortag'))
		# groups = list(Group.objects.values('groupname', 'enrollment_date').distinct().order_by('-enrollment_date'))
		groups = list(Group.objects.select_related('idayear').values('groupname', 'idayear__title').distinct().order_by('-idayear__title'))
		
		# for group in groups:
		# 	group['enrollment_date'] = str(group['enrollment_date'].year) + ' \ ' \
		# 					+ str(group['enrollment_date'].year + 1)
			
		return render(request, 'dashboard_error_groups.html', {'right': True, 'languages': languages, 'tags': tags,
								       'groups': groups})
	else:
		list_filters = json.loads(request.body)
		text = list_filters['text']
		text_type = list_filters['text_type']
		group = list_filters['group']
		tag = list_filters['tag']
		checked_tag_children = list_filters['checked_tag_children']
		
		tags = [tag]
		if checked_tag_children:
			tags = dashboards.get_tag_children(tag)
			
		group_number = []
		group_date = []
		for group in group:
			idx = group.find("(")
			number = int(group[:idx])
			group_number.append(number)
			
			year = group[idx + 2:idx + 6]
			date = year + '-09-01'
			group_date.append(date)
			
		data = []
		texts = []
		text_types = []
		
		for i in range(len(group_number)):
			if group and text and tag and text_type:
				d = list(Error.objects.annotate(id_group=F('sentence__text_id__tbltextgroup__group__id_group'),
								    number=F('sentence__text_id__tbltextgroup__group__group_name'),
								    date=F(
									    'sentence__text_id__tbltextgroup__group__enrollment_date')).values(
					'id_group', 'number', 'date', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
					Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(number=group_number[i]) & Q(
						date=group_date[i]) & Q(tag__id_tag__in=tags) & Q(sentence__text_id__header=text) & Q(
						sentence__text_id__text_type=text_type)).annotate(count_data=Count('id_group')))
				
			elif group and tag and text:
				d = list(Error.objects.annotate(id_group=F('sentence__text_id__tbltextgroup__group__id_group'),
								    number=F('sentence__text_id__tbltextgroup__group__group_name'),
								    date=F(
									    'sentence__text_id__tbltextgroup__group__enrollment_date')).values(
					'id_group', 'number', 'date', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
					Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(number=group_number[i]) & Q(
						date=group_date[i]) & Q(tag__id_tag__in=tags) & Q(sentence__text_id__header=text)).annotate(
					count_data=Count('id_group')))
				
			elif group and tag and text_type:
				d = list(Error.objects.annotate(id_group=F('sentence__text_id__tbltextgroup__group__id_group'),
								    number=F('sentence__text_id__tbltextgroup__group__group_name'),
								    date=F(
									    'sentence__text_id__tbltextgroup__group__enrollment_date')).values(
					'id_group', 'number', 'date', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
					Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(number=group_number[i]) & Q(
						date=group_date[i]) & Q(tag__id_tag__in=tags) & Q(
						sentence__text_id__text_type=text_type)).annotate(count_data=Count('id_group')))
				
			elif group and tag:
				d = list(Error.objects.annotate(id_group=F('sentence__text_id__tbltextgroup__group__id_group'),
								    number=F('sentence__text_id__tbltextgroup__group__group_name'),
								    date=F(
									    'sentence__text_id__tbltextgroup__group__enrollment_date')).values(
					'id_group', 'number', 'date', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
					Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(number=group_number[i]) & Q(
						date=group_date[i]) & Q(tag__id_tag__in=tags)).annotate(count_data=Count('id_group')))
				
			if d != []:
				d = dashboards.get_data_on_tokens(d, '', False, True)
			else:
				d = list(Group.objects.annotate(tag__tag_language=F('language'), number=F('groupname'),
								   date=F('enrollment_date')).values('tag__tag_language',
												     'id_group', 'number',
												     'date').filter(
					Q(number=group_number[i]) & Q(date=group_date[i])))
				
				d[0]['count_data'] = 0
				d[0]['count_data_on_tokens'] = 0
				
			data.append(d)
			
			if text:
				text_types_for_group = list(TextType.objects.values().filter(Q(text__errorcheckflag=True) & Q(
					tbltext__tbltextgroup__group__group_name=group_number[i]) & Q(
					tbltext__tbltextgroup__group__enrollment_date=group_date[i]) & Q(
					tbltext__tblsentence__tblmarkup__tag__in=tags) & Q(tbltext__header=text)).distinct().order_by(
					'idtexttype'))
			else:
				text_types_for_group = list(TextType.objects.values().filter(Q(text__errorcheckflag=True) & Q(
					tbltext__tbltextgroup__group__group_name=group_number[i]) & Q(
					tbltext__tbltextgroup__group__enrollment_date=group_date[i]) & Q(
					tbltext__tblsentence__tblmarkup__tag__in=tags)).distinct().order_by('idtexttype'))
				
			for type_text in text_types_for_group:
				if type_text not in text_types:
					text_types.append(type_text)
					
			if text_type:
				texts_for_group = list(Text.objects.values('header', 'language').filter(
					Q(errorcheckflag=True) & Q(tbltextgroup__group__group_name=group_number[i]) & Q(
						tbltextgroup__group__enrollment_date=group_date[i]) & Q(text_type=text_type) & Q(
						tblsentence__tblmarkup__tag__in=tags)).distinct().order_by('header'))
			else:
				texts_for_group = list(Text.objects.values('header', 'language').filter(
					Q(errorcheckflag=True) & Q(tbltextgroup__group__group_name=group_number[i]) & Q(
						tbltextgroup__group__enrollment_date=group_date[i]) & Q(
						tblsentence__tblmarkup__tag__in=tags)).distinct().order_by('header'))
				
			for text_for_group in texts_for_group:
				if text_for_group not in texts:
					texts.append(text_for_group)
					
		data_all = []
		for i in range(len(data)):
			for data_item in data[i]:
				data_item['date'] = str(data_item['date'].year) + ' \ ' \
							+ str(data_item['date'].year + 1)
				data_all.append(data_item)
				
		data_all = sorted(data_all, key=lambda d: d['count_data'], reverse=True)
		
		return JsonResponse({'data': data_all, 'texts': texts, 'text_types': text_types}, status=200)


def chart_emotions_errors(request):		
	if request.method != 'POST':
		languages = ['Deustache']
		levels = dashboards.get_levels()
		emotions = list(Emotion.objects.values())
		tag_parents, dict_children = dashboards.get_dict_children()
		
		return render(request, 'dashboard_error_emotions.html', {'right': True, 'languages': languages,
									 'levels': levels, 'emotions': emotions,
									 'tag_parents': tag_parents,
									 'dict_children': dict_children})
	else:
		list_filters = json.loads(request.body)
		flag_post = list_filters['flag_post']
		
		if flag_post == 'enrollment_date':
			enrollment_date = dashboards.get_enrollment_date(list_filters)
			return JsonResponse({'enrollment_date': enrollment_date}, status=200)
			
		if flag_post == 'update_diagrams':
			group = list_filters['group']
			date = list_filters['enrollment_date']
			surname = list_filters['surname']
			name = list_filters['name']
			patronymic = list_filters['patronymic']
			course = list_filters['course']
			text = list_filters['text']
			text_type = list_filters['text_type']
			emotion = list_filters['emotion']
			level = int(list_filters['level'])
			
			data_count_errors = []
			if surname and name and patronymic and text and text_type and emotion:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__emotional=emotion) & Q(sentence__text_id__user__last_name=surname) & Q(
							sentence__text_id__user__name=name) & Q(sentence__text_id__user__patronymic=patronymic) & Q(
							sentence__text_id__header=text) & Q(sentence__text_id__text_type=text_type)).annotate(
						count_data=Count('iderrortag__iderrortag')))
				
			elif surname and name and patronymic and text and emotion:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__emotional=emotion) & Q(sentence__text_id__user__last_name=surname) & Q(
							sentence__text_id__user__name=name) & Q(sentence__text_id__user__patronymic=patronymic) & Q(
							sentence__text_id__header=text)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif surname and name and patronymic and text_type and emotion:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__emotional=emotion) & Q(sentence__text_id__user__last_name=surname) & Q(
							sentence__text_id__user__name=name) & Q(sentence__text_id__user__patronymic=patronymic) & Q(
							sentence__text_id__text_type=text_type)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif surname and name and patronymic and emotion:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__emotional=emotion) & Q(sentence__text_id__user__last_name=surname) & Q(
							sentence__text_id__user__name=name) & Q(
							sentence__text_id__user__patronymic=patronymic)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif surname and name and text and text_type and emotion:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__emotional=emotion) & Q(sentence__text_id__user__last_name=surname) & Q(
							sentence__text_id__user__name=name) & Q(sentence__text_id__header=text) & Q(
							sentence__text_id__text_type=text_type)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif surname and name and text and emotion:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__emotional=emotion) & Q(sentence__text_id__user__last_name=surname) & Q(
							sentence__text_id__user__name=name) & Q(sentence__text_id__header=text)).annotate(
						count_data=Count('iderrortag__iderrortag')))
				
			elif surname and name and text_type and emotion:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__emotional=emotion) & Q(sentence__text_id__user__last_name=surname) & Q(
							sentence__text_id__user__name=name) & Q(sentence__text_id__text_type=text_type)).annotate(
						count_data=Count('iderrortag__iderrortag')))
				
			elif surname and name and emotion:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__emotional=emotion) & Q(sentence__text_id__user__last_name=surname) & Q(
							sentence__text_id__user__name=name)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif course and text_type and text and emotion:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__emotional=emotion) & Q(
							sentence__text_id__tbltextgroup__group__course_number=course) & Q(
							sentence__text_id__header=text) & Q(sentence__text_id__text_type=text_type)).annotate(
						count_data=Count('iderrortag__iderrortag')))
				
			elif course and text and emotion:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__emotional=emotion) & Q(
							sentence__text_id__tbltextgroup__group__course_number=course) & Q(
							sentence__text_id__header=text)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif course and text_type and emotion:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__emotional=emotion) & Q(
							sentence__text_id__tbltextgroup__group__course_number=course) & Q(
							sentence__text_id__text_type=text_type)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif course and emotion:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__emotional=emotion) & Q(
							sentence__text_id__tbltextgroup__group__course_number=course)).annotate(
						count_data=Count('iderrortag__iderrortag')))
				
			elif group and text and text_type and emotion:
				group_date = date[:4] + '-09-01'
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__emotional=emotion) & Q(
							sentence__text_id__tbltextgroup__group__group_name=group) & Q(
							sentence__text_id__tbltextgroup__group__enrollment_date=group_date) & Q(
							sentence__text_id__header=text) & Q(sentence__text_id__text_type=text_type)).annotate(
						count_data=Count('iderrortag__iderrortag')))
				
			elif group and text and emotion:
				group_date = date[:4] + '-09-01'
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__emotional=emotion) & Q(
							sentence__text_id__tbltextgroup__group__group_name=group) & Q(
							sentence__text_id__tbltextgroup__group__enrollment_date=group_date) & Q(
							sentence__text_id__header=text)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif group and text_type and emotion:
				group_date = date[:4] + '-09-01'
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__emotional=emotion) & Q(
							sentence__text_id__tbltextgroup__group__group_name=group) & Q(
							sentence__text_id__tbltextgroup__group__enrollment_date=group_date) & Q(
							sentence__text_id__text_type=text_type)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif group and emotion:
				group_date = date[:4] + '-09-01'
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__emotional=emotion) & Q(
							sentence__text_id__tbltextgroup__group__group_name=group) & Q(
							sentence__text_id__tbltextgroup__group__enrollment_date=group_date)).annotate(
						count_data=Count('iderrortag__iderrortag')))
				
			elif text_type and text and emotion:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__emotional=emotion) & Q(sentence__text_id__header=text) & Q(
							sentence__text_id__text_type=text_type)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif text_type and emotion:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__emotional=emotion) & Q(sentence__text_id__text_type=text_type)).annotate(
						count_data=Count('iderrortag__iderrortag')))
				
			elif text and emotion:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__emotional=emotion) & Q(sentence__text_id__header=text)).annotate(
						count_data=Count('iderrortag__iderrortag')))
				
			elif emotion:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__emotional=emotion)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			data_count_on_tokens = dashboards.get_data_on_tokens(data_count_errors, 'iderrortag__iderrortag', 'tag__tag_language',
									     True, False)
			data_errors = dashboards.get_data_errors(data_count_on_tokens, level, True)
			
			groups, courses, texts = dashboards.get_filters_for_choice_text_type(list_filters)
			_, _, text_types = dashboards.get_filters_for_choice_text(list_filters)
			
			return JsonResponse({'data': data_errors, 'groups': groups, 'courses': courses, 'texts': texts,
					     'text_types': text_types}, status=200)


def chart_self_rating_errors(request):	
	if request.method != 'POST':
		languages = ['Deustache']
		levels = dashboards.get_levels()
		self_ratings = list(Text.objects.values('selfrating').filter(
			Q(selfrating__gt=0) & Q(errorcheckflag=True)).distinct().order_by('selfrating'))
		tag_parents, dict_children = dashboards.get_dict_children()
		
		self_rating_text = Text.TASK_RATES
		
		for selfrating in self_ratings:
			idx = selfrating["selfrating"]
			selfrating["self_rating_text"] = self_rating_text[idx - 1][1]
			
		return render(request, 'dashboard_error_self_rating.html', {'right': True, 'languages': languages,
									    'levels': levels, 'tag_parents': tag_parents,
									    'self_ratings': self_ratings,
									    'dict_children': dict_children})
	else:
		list_filters = json.loads(request.body)
		flag_post = list_filters['flag_post']
		
		if flag_post == 'enrollment_date':
			enrollment_date = dashboards.get_enrollment_date(list_filters)
			return JsonResponse({'enrollment_date': enrollment_date}, status=200)
			
		if flag_post == 'update_diagrams':
			group = list_filters['group']
			date = list_filters['enrollment_date']
			surname = list_filters['surname']
			name = list_filters['name']
			patronymic = list_filters['patronymic']
			course = list_filters['course']
			text = list_filters['text']
			text_type = list_filters['text_type']
			selfrating = list_filters['selfrating']
			level = int(list_filters['level'])
			
			data_count_errors = []
			if surname and name and patronymic and text and text_type and selfrating:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__self_rating=selfrating) & Q(
							sentence__text_id__user__last_name=surname) & Q(sentence__text_id__user__name=name) & Q(
							sentence__text_id__user__patronymic=patronymic) & Q(sentence__text_id__header=text) & Q(
							sentence__text_id__text_type=text_type)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif surname and name and patronymic and text and selfrating:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__self_rating=selfrating) & Q(
							sentence__text_id__user__last_name=surname) & Q(sentence__text_id__user__name=name) & Q(
							sentence__text_id__user__patronymic=patronymic) & Q(
							sentence__text_id__header=text)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif surname and name and patronymic and text_type and selfrating:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__self_rating=selfrating) & Q(
							sentence__text_id__user__last_name=surname) & Q(sentence__text_id__user__name=name) & Q(
							sentence__text_id__user__patronymic=patronymic) & Q(
							sentence__text_id__text_type=text_type)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif surname and name and patronymic and selfrating:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__self_rating=selfrating) & Q(
							sentence__text_id__user__last_name=surname) & Q(sentence__text_id__user__name=name) & Q(
							sentence__text_id__user__patronymic=patronymic)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif surname and name and text and text_type and selfrating:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__self_rating=selfrating) & Q(
							sentence__text_id__user__last_name=surname) & Q(sentence__text_id__user__name=name) & Q(
							sentence__text_id__header=text) & Q(sentence__text_id__text_type=text_type)).annotate(
						count_data=Count('iderrortag__iderrortag')))
				
			elif surname and name and text and selfrating:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__self_rating=selfrating) & Q(
							sentence__text_id__user__last_name=surname) & Q(sentence__text_id__user__name=name) & Q(
							sentence__text_id__header=text)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif surname and name and text_type and selfrating:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__self_rating=selfrating) & Q(
							sentence__text_id__user__last_name=surname) & Q(sentence__text_id__user__name=name) & Q(
							sentence__text_id__text_type=text_type)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif surname and name and selfrating:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__self_rating=selfrating) & Q(
							sentence__text_id__user__last_name=surname) & Q(
							sentence__text_id__user__name=name)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif course and text_type and text and selfrating:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__self_rating=selfrating) & Q(
							sentence__text_id__tbltextgroup__group__course_number=course) & Q(
							sentence__text_id__header=text) & Q(sentence__text_id__text_type=text_type)).annotate(
						count_data=Count('iderrortag__iderrortag')))
				
			elif course and text and selfrating:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__self_rating=selfrating) & Q(
							sentence__text_id__tbltextgroup__group__course_number=course) & Q(
							sentence__text_id__header=text)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif course and text_type and selfrating:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__self_rating=selfrating) & Q(
							sentence__text_id__tbltextgroup__group__course_number=course) & Q(
							sentence__text_id__text_type=text_type)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif course and selfrating:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__self_rating=selfrating) & Q(
							sentence__text_id__tbltextgroup__group__course_number=course)).annotate(
						count_data=Count('iderrortag__iderrortag')))
				
			elif group and text and text_type and selfrating:
				group_date = date[:4] + '-09-01'
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__self_rating=selfrating) & Q(
							sentence__text_id__tbltextgroup__group__group_name=group) & Q(
							sentence__text_id__tbltextgroup__group__enrollment_date=group_date) & Q(
							sentence__text_id__header=text) & Q(sentence__text_id__text_type=text_type)).annotate(
						count_data=Count('iderrortag__iderrortag')))
				
			elif group and text and selfrating:
				group_date = date[:4] + '-09-01'
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__self_rating=selfrating) & Q(
							sentence__text_id__tbltextgroup__group__group_name=group) & Q(
							sentence__text_id__tbltextgroup__group__enrollment_date=group_date) & Q(
							sentence__text_id__header=text)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif group and text_type and selfrating:
				group_date = date[:4] + '-09-01'
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__self_rating=selfrating) & Q(
							sentence__text_id__tbltextgroup__group__group_name=group) & Q(
							sentence__text_id__tbltextgroup__group__enrollment_date=group_date) & Q(
							sentence__text_id__text_type=text_type)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif group and selfrating:
				group_date = date[:4] + '-09-01'
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__self_rating=selfrating) & Q(
							sentence__text_id__tbltextgroup__group__group_name=group) & Q(
							sentence__text_id__tbltextgroup__group__enrollment_date=group_date)).annotate(
						count_data=Count('iderrortag__iderrortag')))
				
			elif text_type and text and selfrating:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__self_rating=selfrating) & Q(sentence__text_id__header=text) & Q(
							sentence__text_id__text_type=text_type)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif text_type and selfrating:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__self_rating=selfrating) & Q(
							sentence__text_id__text_type=text_type)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			elif text and selfrating:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__self_rating=selfrating) & Q(sentence__text_id__header=text)).annotate(
						count_data=Count('iderrortag__iderrortag')))
				
			elif selfrating:
				data_count_errors = list(
					Error.objects.values('iderrortag__iderrortag', 'iderrortag__idtagparent', 'iderrortag__tagtext',
								 'iderrortag__tagtextrussian', 'errortoken__idtoken__idsentence__idtext__idtext').filter(
						Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
							sentence__text_id__self_rating=selfrating)).annotate(count_data=Count('iderrortag__iderrortag')))
				
			data_count_on_tokens = dashboards.get_data_on_tokens(data_count_errors, 'iderrortag__iderrortag', 'tag__tag_language',
									     True, False)
			data_errors = dashboards.get_data_errors(data_count_on_tokens, level, True)
			
			groups, courses, texts = dashboards.get_filters_for_choice_text_type(list_filters)
			_, _, text_types = dashboards.get_filters_for_choice_text(list_filters)
			
			return JsonResponse({'data': data_errors, 'groups': groups, 'courses': courses, 'texts': texts,
					     'text_types': text_types}, status=200)


def chart_relation_assessment_self_rating(request):	
	if request.method != 'POST':
		languages = ['Deustache']
		
		return render(request, 'dashboard_assessment_self_rating.html', {'right': True, 'languages': languages})
	else:
		list_filters = json.loads(request.body)
		surname = list_filters['surname']
		name = list_filters['name']
		patronymic = list_filters['patronymic']
		text_type = list_filters['text_type']
		
		if surname and name and patronymic and text_type:
			data_relation = list(Text.objects.values('textgrade', 'selfrating').filter(
				Q(self_rating__gt=0) & Q(assessment__gt=0) & Q(user__last_name=surname) & Q(user__name=name) & Q(
					user__patronymic=patronymic) & Q(text_type=text_type) & Q(errorcheckflag=True)).distinct())
			
		elif surname and name and patronymic:
			data_relation = list(Text.objects.values('textgrade', 'selfrating').filter(
				Q(self_rating__gt=0) & Q(assessment__gt=0) & Q(user__last_name=surname) & Q(user__name=name) & Q(
					user__patronymic=patronymic) & Q(errorcheckflag=True)).distinct())
			
		elif surname and name and text_type:
			data_relation = list(Text.objects.values('textgrade', 'selfrating').filter(
				Q(self_rating__gt=0) & Q(assessment__gt=0) & Q(user__last_name=surname) & Q(user__name=name) & Q(
					text_type=text_type) & Q(errorcheckflag=True)).distinct())
			
		else:
			data_relation = list(Text.objects.values('textgrade', 'selfrating').filter(
				Q(self_rating__gt=0) & Q(assessment__gt=0) & Q(user__last_name=surname) & Q(user__name=name) & Q(
					errorcheckflag=True)).distinct())
			
		assessment_types = Text.TASK_RATES
		
		for data in data_relation:
			idx = data["selfrating"]
			data["self_rating_text"] = assessment_types[idx - 1][1]
			
			idx = data["textgrade"]
			data["assessment_text"] = assessment_types[idx - 1][1]
			
		if patronymic:
			text_types = list(TextType.objects.values().filter(
				Q(tbltext__self_rating__gt=0) & Q(tbltext__assessment__gt=0) & Q(tbltext__user__last_name=surname) & Q(
					tbltext__user__name=name) & Q(tbltext__user__patronymic=patronymic) & Q(
					text__errorcheckflag=True)).distinct().order_by('idtexttype'))
		else:
			text_types = list(TextType.objects.values().filter(
				Q(tbltext__self_rating__gt=0) & Q(tbltext__assessment__gt=0) & Q(tbltext__user__last_name=surname) & Q(
					tbltext__user__name=name) & Q(text__errorcheckflag=True)).distinct().order_by('idtexttype'))
			
		return JsonResponse({'data': data_relation, 'text_types': text_types}, status=200)


# ПОИСК ЗАВИСИМОСТЕЙ

def relation_emotions_self_rating(request):	
	if request.method != 'POST':
		languages = ['Deustache']
		courses = list(
			Group.objects.values('studycourse').filter(studycourse__gt=0).distinct().order_by(
				'studycourse'))
		groups = list(Group.objects.values('groupname').distinct().order_by('groupname'))
		
		data_relation = list(
			Text.objects.values('idemotion', 'selfrating').filter(
				Q(idemotion__isnull=False) & Q(selfrating__gt=0) & ~Q(idemotion=2)))
		
		data, relation, data_fisher = dashboards.get_stat(data_relation, 'idemotion', 'idemotion__emotionname',
								  'selfrating', 'self_rating_text', True)
		
		return render(request, 'relation_emotions_self_rating.html', {'right': True, 'languages': languages,
									      'courses': courses, 'groups': groups,
									      'data_relation': data, 'relation': relation,
									      'data_fisher': data_fisher})
	else:
		list_filters = json.loads(request.body)
		flag_post = list_filters['flag_post']
		
		if flag_post == 'enrollment_date':
			group = list_filters['group']
			enrollment_date = list(
				Group.objects.values('enrollment_date').filter(groupname=group).distinct().order_by(
					'enrollment_date'))
			
			for date in enrollment_date:
				date['enrollment_date'] = str(date['enrollment_date'].year) + ' \ ' \
								+ str(date['enrollment_date'].year + 1)
				
			return JsonResponse({'enrollment_date': enrollment_date}, status=200)
			
		if flag_post == 'course':
			course = list_filters['course']
			
			data_relation = list(Text.objects.values('idemotion', 'selfrating').filter(
				Q(emotional__isnull=False) & Q(self_rating__gt=0) & Q(tbltextgroup__group__course_number=course) & ~Q(
					emotional=2)))
			
		if flag_post == 'group':
			group = list_filters['group']
			date = list_filters['date']
			group_date = date[:4] + '-09-01'
			
			data_relation = list(
				Text.objects.values('idemotion', 'selfrating').filter(
					Q(emotional__isnull=False) & Q(self_rating__gt=0) & Q(
						tbltextgroup__group__group_name=group) & Q(
						tbltextgroup__group__enrollment_date=group_date) & ~Q(emotional=2)))
			
		data, relation, data_fisher = dashboards.get_stat(data_relation, 'idemotion', 'idemotion__emotionname',
								  'selfrating', 'self_rating_text', True)
		
		return JsonResponse({'data_relation': data, 'relation': relation, 'data_fisher': data_fisher}, status=200)


def relation_emotions_assessment(request):	
	if request.method != 'POST':
		languages = ['Deustache']
		courses = list(
			Group.objects.values('studycourse').filter(studycourse__gt=0).distinct().order_by(
				'studycourse'))
		groups = list(Group.objects.values('groupname').distinct().order_by('groupname'))
		
		data_relation = list(
			Text.objects.values('idemotion', 'textgrade').filter(
				Q(idemotion__isnull=False) & Q(textgrade__gt=0) & ~Q(idemotion=2)))
		
		data, relation, data_fisher = dashboards.get_stat(data_relation, 'idemotion', 'idemotion__emotionname',
								  'textgrade', 'assessment_text', True)
		
		return render(request, 'relation_emotions_assessment.html', {'right': True, 'languages': languages,
									     'courses': courses, 'groups': groups,
									     'data_relation': data, 'relation': relation,
									     'data_fisher': data_fisher})
	else:
		list_filters = json.loads(request.body)
		flag_post = list_filters['flag_post']
		
		if flag_post == 'enrollment_date':
			group = list_filters['group']
			enrollment_date = list(
				Group.objects.values('enrollment_date').filter(groupname=group).distinct().order_by(
					'enrollment_date'))
			
			for date in enrollment_date:
				date['enrollment_date'] = str(date['enrollment_date'].year) + ' \ ' \
								+ str(date['enrollment_date'].year + 1)
				
			return JsonResponse({'enrollment_date': enrollment_date}, status=200)
			
		if flag_post == 'course':
			course = list_filters['course']
			
			data_relation = list(
				Text.objects.values('idemotion', 'textgrade').filter(
					Q(emotional__isnull=False) & Q(assessment__gt=0) & Q(
						tbltextgroup__group__course_number=course) & ~Q(emotional=2)))
			
		if flag_post == 'group':
			group = list_filters['group']
			date = list_filters['date']
			group_date = date[:4] + '-09-01'
			
			data_relation = list(Text.objects.values('idemotion', 'textgrade').filter(
				Q(emotional__isnull=False) & Q(assessment__gt=0) & Q(tbltextgroup__group__group_name=group) & Q(
					tbltextgroup__group__enrollment_date=group_date) & ~Q(emotional=2)))
			
		data, relation, data_fisher = dashboards.get_stat(data_relation, 'idemotion', 'idemotion__emotionname',
								  'textgrade', 'assessment_text', True)
		
		return JsonResponse({'data_relation': data, 'relation': relation, 'data_fisher': data_fisher}, status=200)


def relation_self_rating_assessment(request):	
	if request.method != 'POST':
		languages = ['Deustache']
		courses = list(
			Group.objects.values('studycourse').filter(studycourse__gt=0).distinct().order_by(
				'studycourse'))
		groups = list(Group.objects.values('groupname').distinct().order_by('groupname'))
		
		data_relation = list(Text.objects.values('selfrating', 'textgrade').filter(
			Q(selfrating__gt=0) & Q(textgrade__gt=0)))
		
		data, relation, data_fisher = dashboards.get_stat(data_relation, 'selfrating', 'self_rating_text',
								  'textgrade', 'assessment_text', False)
		
		return render(request, 'relation_self_rating_assessment.html', {'right': True, 'languages': languages,
										'courses': courses, 'groups': groups,
										'data_relation': data, 'relation': relation,
										'data_fisher': data_fisher})
	else:
		list_filters = json.loads(request.body)
		flag_post = list_filters['flag_post']
		
		if flag_post == 'enrollment_date':
			group = list_filters['group']
			enrollment_date = list(
				Group.objects.values('enrollment_date').filter(groupname=group).distinct().order_by(
					'enrollment_date'))
			
			for date in enrollment_date:
				date['enrollment_date'] = str(date['enrollment_date'].year) + ' \ ' \
								+ str(date['enrollment_date'].year + 1)
				
			return JsonResponse({'enrollment_date': enrollment_date}, status=200)
			
		if flag_post == 'course':
			course = list_filters['course']
			
			data_relation = list(Text.objects.values('selfrating', 'textgrade').filter(
				Q(self_rating__gt=0) & Q(assessment__gt=0) & Q(tbltextgroup__group__course_number=course)))
			
		if flag_post == 'group':
			group = list_filters['group']
			date = list_filters['date']
			group_date = date[:4] + '-09-01'
			
			data_relation = list(Text.objects.values('selfrating', 'textgrade').filter(
				Q(self_rating__gt=0) & Q(assessment__gt=0) & Q(tbltextgroup__group__group_name=group) & Q(
					tbltextgroup__group__enrollment_date=group_date)))
			
		data, relation, data_fisher = dashboards.get_stat(data_relation, 'selfrating', 'self_rating_text',
								  'textgrade', 'assessment_text', False)
		
		return JsonResponse({'data_relation': data, 'relation': relation, 'data_fisher': data_fisher}, status=200)


def relation_course_errors(request):	
	if request.method != 'POST':
		languages = ['Deustache']
		tags = list(Error.objects.values('iderrortag__iderrortag', 'iderrortag__tagtext', 'iderrortag__tagtextrussian').order_by('iderrortag__iderrortag'))
		
		return render(request, 'relation_course_errors.html', {'right': True, 'languages': languages, 'tags': tags})
	else:
		list_filters = json.loads(request.body)
		flag_post = list_filters['flag_post']
		tag = list_filters['tag']
		checked_tag_children = list_filters['checked_tag_children']
		
		tags = [tag]
		if checked_tag_children:
			tags = dashboards.get_tag_children(tag)
			
		if flag_post == 'courses':
			data_relation = list(
				Error.objects.values('sentence__text_id__tbltextgroup__group__course_number').filter(
					Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(tag__id_tag__in=tags) & Q(
						sentence__text_id__tbltextgroup__group__course_number__isnull=False)).annotate(
					count_data=Count('sentence__text_id__tbltextgroup__group__course_number')))
			
		if flag_post == 'students':
			data_relation = list(Error.objects.values('sentence__text_id__user',
								      'sentence__text_id__tbltextgroup__group__course_number').filter(
				Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(tag__id_tag__in=tags) & Q(
					sentence__text_id__tbltextgroup__group__course_number__isnull=False)).annotate(
				count_data=Count('sentence__text_id__tbltextgroup__group__course_number')))
			
		if flag_post == 'groups':
			data_relation = list(Error.objects.values('sentence__text_id__tbltextgroup__group',
								      'sentence__text_id__tbltextgroup__group__course_number').filter(
				Q(tag__markup_type=1) & Q(sentence__text_id__error_tag_check=1) & Q(
					sentence__text_id__tbltextgroup__group__isnull=False) & Q(tag__id_tag__in=tags)).annotate(
				count_data=Count('sentence__text_id__tbltextgroup__group__course_number')))
			
		course = []
		count_errors = []
		
		for data in data_relation:
			course.append(data['sentence__text_id__tbltextgroup__group__course_number'])
			count_errors.append(data['count_data'])
			
		critical_stat_level = 0.05
		n = len(course)
		
		if n > 1:
			if len(set(course)) == 1 or len(set(count_errors)) == 1:
				relation = {'result': 'один из параметров константа', 'stat': 'None', 'pvalue': 'None', 'N': n}
				
			else:
				result = scipy.stats.spearmanr(course, count_errors)
				
				t = abs(result.statistic) * np.sqrt((n-2) / (1 - result.statistic * result.statistic))
				t_critical = scipy.stats.t.ppf(1-critical_stat_level/2, n-2)
				
				if t < t_critical:
					worth = 'корреляция статистически не значимая'
				else:
					worth = 'статистически значимая корреляция'
					
				if np.isnan(result.pvalue):
					pvalue = 'Nan'
				else:
					pvalue = result.pvalue
					
				if result.statistic == 0:
					relation = {'result': f'связь отсутствует  ({worth})', 'stat': result.statistic, 'pvalue': pvalue,
						    'N': n}
				elif result.statistic >= 0.75:
					relation = {'result': f'очень высокая положительная связь ({worth})', 'stat': result.statistic,
						    'pvalue': pvalue, 'N': n}
				elif 0.5 <= result.statistic < 0.75:
					relation = {'result': f'высокая положительная связь  ({worth})', 'stat': result.statistic,
						    'pvalue': pvalue, 'N': n}
				elif 0.25 <= result.statistic < 0.5:
					relation = {'result': f'средняя положительная связь  ({worth})', 'stat': result.statistic,
						    'pvalue': pvalue, 'N': n}
				elif 0 < result.statistic < 0.25:
					relation = {'result': f'слабая положительная связь  ({worth})', 'stat': result.statistic,
						    'pvalue': pvalue, 'N': n}
				elif -0.25 <= result.statistic < 0:
					relation = {'result': f'слабая отрицательная связь  ({worth})', 'stat': result.statistic,
						    'pvalue': pvalue, 'N': n}
				elif -0.5 <= result.statistic < -0.25:
					relation = {'result': f'средняя отрицательная связь  ({worth})', 'stat': result.statistic,
						    'pvalue': pvalue, 'N': n}
				elif -0.75 <= result.statistic < -0.5:
					relation = {'result': f'высокая отрицательная связь  ({worth})', 'stat': result.statistic,
						    'pvalue': pvalue, 'N': n}
				else:
					relation = {'result': f'очень высокая отрицательная связь  ({worth})', 'stat': result.statistic,
						    'pvalue': pvalue, 'N': n}
		else:
			relation = {'result': '-', 'stat': '-', 'pvalue': '-', 'N': n}
			
		return JsonResponse({'data_relation': data_relation, 'relation': relation}, status=200)